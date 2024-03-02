#import arcpy
import os
import pandas as pd
import geopandas as gpd
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
from database.tradutor import Tradutor
import arcpy
import shutil
arcpy.env.overwriteOutput = True

class XLSXConversion:
    def __init__(self, feature_class, related_field, excel_file, table_description, category):
        self.feature_class = feature_class
        self.related_field = related_field
        self.excel_file = excel_file
        self.feature_class_name = os.path.basename(feature_class)
        self.table_description = table_description

    def compare_fields(self):
        campos_fc = [campo.name for campo in arcpy.ListFields(self.feature_class)]
        colunas_comuns = [campo for campo in campos_fc if campo in self.related_field]
        return colunas_comuns

    def read_and_import_dataframe(self):
        arc = os.path.join(os.path.dirname(self.excel_file),'gpd_read_files')
        os.makedirs(arc, exist_ok=True)

        arcpy.conversion.FeatureClassToShapefile(self.feature_class, arc)
        feature_class = os.path.join(arc, os.path.basename(self.feature_class))+'.shp'
        df = gpd.read_file(feature_class)
        df = df[self.compare_fields()]
        
        workbook = op.load_workbook(self.excel_file)
        sheet = workbook.create_sheet(title=self.feature_class_name.split('__')[0])

        for r in dataframe_to_rows(df, index = False, header = True):
            sheet.append(r)

        #apaga todas as colunas que não tem dados
        for col in range(1, sheet.max_column+1):
            if sheet.cell(row=1, column=col).value == None:
                sheet.delete_cols(col)

        if sheet.cell(row=3, column=1).value == None:
            self.empty_sheet(sheet)

        self.style_sheet(sheet)
        tradutor = Tradutor(sheet)
        tradutor.tradutor()
        
        workbook.save(self.excel_file)
    

    def style_sheet(self, sheet):
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = op.styles.Font(bold=True)
                cell.fill = op.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                cell.border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                  right=op.styles.Side(border_style="thin"),
                                                  top=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="thin"))
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                  right=op.styles.Side(border_style="thin"),
                                                  top=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="thin"))
        for col in range(1, sheet.max_column+1):
            sheet.column_dimensions[op.utils.get_column_letter(col)].width = 30
        sheet.insert_rows(1)
        sheet.cell(row=1, column=1, value=self.table_description)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
        sheet.cell(row=1, column=1).fill = op.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
        sheet.cell(row=1, column=1).alignment = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.cell(row=1, column=1).font = op.styles.Font(bold=True, size=14)
        sheet.row_dimensions[1].height = 50
        sheet.cell(row=1, column=1).value = sheet.cell(row=1, column=1).value.replace("_", " ")
        

    def empty_sheet(self, sheet):
        sheet.insert_rows(2)
        sheet.cell(row=2, column=1, value="Não há dados do tema na área de estudo")
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=sheet.max_column)
        sheet.cell(row=2, column=1).alignment = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    def get_complete_sheet(self):
        self.compare_fields()
        self.read_and_import_dataframe()
        arc = os.path.join(os.path.dirname(self.excel_file),'gpd_read_files')
        shutil.rmtree(arc)
    
        

        



